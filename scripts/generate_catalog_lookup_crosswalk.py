from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import range_boundaries, get_column_letter


WORKBOOK_PATH = Path("catelog_lookup.xlsx")
UPDATED_WORKBOOK_PATH = Path("catelog_lookup_with_keys.xlsx")
CROSSWALK_CSV_PATH = Path("catalog_lookup_crosswalk.csv")
TABLE1_NAME = "Table1"  # new catalog
TABLE2_NAME = "Table2"  # old catalog
UNIQUE_KEY_HEADER = "UniqueKey"


@dataclass
class TableRow:
    row_index: int
    values: Dict[str, object]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_unique_key(cost_code: object, item_id: object) -> str:
    return f"{normalize_text(cost_code)}|{normalize_text(item_id)}"


def read_table(ws: Worksheet, table: Table) -> Tuple[List[str], List[TableRow], Tuple[int, int, int, int]]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [normalize_text(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]

    rows: List[TableRow] = []
    for r in range(min_row + 1, max_row + 1):
        values = {
            headers[i]: ws.cell(r, min_col + i).value for i in range(len(headers))
        }
        rows.append(TableRow(row_index=r, values=values))

    return headers, rows, (min_col, min_row, max_col, max_row)


def ensure_unique_key_column(
    ws: Worksheet,
    table: Table,
    headers: List[str],
    bounds: Tuple[int, int, int, int],
    rows: List[TableRow],
) -> None:
    min_col, min_row, max_col, max_row = bounds

    if UNIQUE_KEY_HEADER in headers:
        key_col = min_col + headers.index(UNIQUE_KEY_HEADER)
    else:
        key_col = max_col + 1
        ws.cell(min_row, key_col, UNIQUE_KEY_HEADER)
        max_col = key_col
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    for row in rows:
        cost_code = row.values.get("Cost Code")
        item_id = row.values.get("ItemId")
        ws.cell(row.row_index, key_col, build_unique_key(cost_code, item_id))


def write_crosswalk(old_rows: List[TableRow], new_rows: List[TableRow]) -> None:
    total = min(len(old_rows), len(new_rows))

    with CROSSWALK_CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "RowIndex",
                "OldUniqueKey",
                "OldCostCode",
                "OldItemId",
                "OldName",
                "NewUniqueKey",
                "NewCostCode",
                "NewItemId",
                "NewName",
            ]
        )

        for i in range(total):
            old = old_rows[i].values
            new = new_rows[i].values
            writer.writerow(
                [
                    i + 1,
                    build_unique_key(old.get("Cost Code"), old.get("ItemId")),
                    normalize_text(old.get("Cost Code")),
                    normalize_text(old.get("ItemId")),
                    normalize_text(old.get("Name")),
                    build_unique_key(new.get("Cost Code"), new.get("ItemId")),
                    normalize_text(new.get("Cost Code")),
                    normalize_text(new.get("ItemId")),
                    normalize_text(new.get("Name")),
                ]
            )


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Sheet1"]

    table1 = ws.tables[TABLE1_NAME]
    table2 = ws.tables[TABLE2_NAME]

    t1_headers, t1_rows, t1_bounds = read_table(ws, table1)
    t2_headers, t2_rows, t2_bounds = read_table(ws, table2)

    ensure_unique_key_column(ws, table1, t1_headers, t1_bounds, t1_rows)
    ensure_unique_key_column(ws, table2, t2_headers, t2_bounds, t2_rows)

    wb.save(UPDATED_WORKBOOK_PATH)
    write_crosswalk(t2_rows, t1_rows)

    print(f"Updated workbook: {UPDATED_WORKBOOK_PATH}")
    print(f"Crosswalk CSV: {CROSSWALK_CSV_PATH}")
    print(f"Rows mapped: {min(len(t1_rows), len(t2_rows))}")


if __name__ == "__main__":
    main()
